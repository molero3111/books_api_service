from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from apps.authors.models import Author
from apps.books.models import Book


def export_authors_books(request):
    # Create a workbook and sheets
    wb = Workbook()
    ws_authors = wb.active
    ws_authors.title = 'Authors'
    ws_books = wb.create_sheet(title='Books')

    # Define the columns for the Authors sheet
    author_columns = ['ID', 'User ID', 'Nickname', 'Published Books', 'Created At', 'Updated At']
    for col_num, column_title in enumerate(author_columns, 1):
        col_letter = get_column_letter(col_num)
        ws_authors[f'{col_letter}1'] = column_title
        ws_authors[f'{col_letter}1'].alignment = Alignment(horizontal='center')

    # Add data to Authors sheet
    authors = Author.objects.all()
    for row_num, author in enumerate(authors, 2):
        ws_authors[f'A{row_num}'] = author.id
        ws_authors[f'B{row_num}'] = author.user_id
        ws_authors[f'C{row_num}'] = author.nickname
        ws_authors[f'D{row_num}'] = author.published_books
        ws_authors[f'E{row_num}'] = author.created_at.strftime('%Y-%m-%d %H:%M:%S') if author.created_at else ''
        ws_authors[f'F{row_num}'] = author.updated_at.strftime('%Y-%m-%d %H:%M:%S') if author.updated_at else ''

    # Define the columns for the Books sheet
    book_columns = ['ID', 'Author ID', 'Title', 'Genre', 'Published At', 'Created At', 'Updated At']
    for col_num, column_title in enumerate(book_columns, 1):
        col_letter = get_column_letter(col_num)
        ws_books[f'{col_letter}1'] = column_title
        ws_books[f'{col_letter}1'].alignment = Alignment(horizontal='center')

    # Add data to Books sheet
    books = Book.objects.all()
    for row_num, book in enumerate(books, 2):
        ws_books[f'A{row_num}'] = book.id
        ws_books[f'B{row_num}'] = book.author_id
        ws_books[f'C{row_num}'] = book.title
        ws_books[f'D{row_num}'] = book.genre
        ws_books[f'E{row_num}'] = book.published_at.strftime('%Y-%m-%d %H:%M:%S') if book.published_at else ''
        ws_books[f'F{row_num}'] = book.created_at.strftime('%Y-%m-%d %H:%M:%S') if book.created_at else ''
        ws_books[f'G{row_num}'] = book.updated_at.strftime('%Y-%m-%d %H:%M:%S') if book.updated_at else ''

    # Save the workbook to a bytes buffer
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=authors_books.xlsx'
    wb.save(response)

    return response
